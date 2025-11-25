import requests
from openpyxl import load_workbook
from exel_list import write_to_sheet_1, sheet

arr = []
def find_item(response, id_articul: int):
    date = response.json()
    products = date['products']
    for index, value in enumerate(products):
        if int(value['id']) == id_articul:
            return index + 1

    return None

def get_resp():

    # Открываем файл
    #wb = load_workbook("shorts_position_words.xlsx")

    # Выбираем активный лист
    #ws = wb.active
    exel = sheet.col_values(1)
    # Считываем данные из столбца J (10-й столбец)
    #keywords = [cell.value for cell in ws["A"] if cell.value is not None]
    keywords = [cell for cell in exel if cell]

    for keyword in keywords:
        page = 1
        while True:
            cookies = {
                'external-locale': 'ru',
                '_wbauid': '8264840631760617509',
                'wbx-validation-key': '0b839a54-4dfe-4b2a-a0a6-edf2ba0b6db4',
                'x-supplier-id-external': '0cd68178-a8ec-409f-8817-787cd8da5b16',
                '_cp': '1',
                'x_wbaas_token': '1.1000.8fe21158a99145ff97177af0a0a36575.MHw3Ny4yNDMuMjkuMTI2fE1vemlsbGEvNS4wIChXaW5kb3dzIE5UIDEwLjA7IFdpbjY0OyB4NjQpIEFwcGxlV2ViS2l0LzUzNy4zNiAoS0hUTUwsIGxpa2UgR2Vja28pIENocm9tZS8xNDAuMC4wLjAgWWFCcm93c2VyLzI1LjEwLjAuMCBTYWZhcmkvNTM3LjM2fDE3NjUyNjQ4MzN8cmV1c2FibGV8MnxleUpvWVhOb0lqb2lJbjA9fDB8M3wxNzY0NjYwMDMzfDE=.MEUCIQDdvJ69jwv5y4qD8HkfWPfE7NttWzAApDZDDXLrC9RQ3AIgWQJ8j9JsEFOgO9iAV1IKCisj99mKd+a6TQweuIPwpnI=',
                '__zzatw-wb': 'MDA0dC0cTHtmcDhhDHEWTT17CT4VHThHKHIzd2UqP2YlaE5eJjVRP0FaW1Q4NmdBEXUmCQg3LGBwVxlRExpceEdXeiwfFXdwKFIKEF5CSmllbQwtUlFRS19/Dg4/aU5ZQ11wS3E6EmBWGB5CWgtMeFtLKRZHGzJhXkZpdRVSCxFeQnN0dCw9bFBjShVQdBAICFpNRXkmKk8NFF1BR19vG3siXyoIJGM1Xz9EaVhTMCpYQXt1J3Z+KmUzPGwjY0deJEZYUnwrGw1pN2wXPHVlLwkxLGJ5MVIvE0tsP0caRFpbQDsyVghDQE1HFF9BWncyUlFRS2EQR0lrZU5TQixmG3EVTQgNND1aciIPWzklWAgSPwsmIBd7ayhUCg5gQEdxbxt/Nl0cOWMRCxl+OmNdRkc3FSR7dSYKCTU3YnAvTCB7SykWRxsyYV5GaXUVTzo/YUVCdHsmbG1SGkRdH0wTSgktGhh0citWOj9jcXJyLSpBVxlRDxZhDhYYRRcje0I3Yhk4QhgvPV8/YngiD2lIYCVHVVJ+KBwWe3EmS3FPLH12X30beylOIA0lVBMhP05yIeFnyg==',
                'cfidsw-wb': 'XsAr4MnOWjEwlWh4U96Tq+AFP3MrXypo1QoN28dMc9pZnWZAHDjH0LSKjAyYELUerfy2t40DHt6gh+ARWmZoUTCf+tOrRHCIzokKcqffeGBN7zBQ+wNC6vG1cRQY/0QpTeyzO2X+yBSLAvZbHoaxbQlUi1ODq9ReFSe7ewQ=',
            }

            headers = {
                'accept': '*/*',
                'accept-language': 'ru,en;q=0.9',
                'deviceid': 'site_9e75b73088c84aa39d1db10f8b866328',
                'priority': 'u=1, i',
                'referer': 'https://www.wildberries.ru/catalog/0/search.aspx?search=%D0%BA%D0%B0%D0%BB%D1%8C%D1%81%D0%BE%D0%BD%D1%8B+%D0%B4%D0%BB%D1%8F+%D0%BC%D0%B0%D0%BB%D1%8C%D1%87%D0%B8%D0%BA%D0%B0',
                'sec-ch-ua': '"Chromium";v="140", "Not=A?Brand";v="24", "YaBrowser";v="25.10", "Yowser";v="2.5"',
                'sec-ch-ua-mobile': '?0',
                'sec-ch-ua-platform': '"Windows"',
                'sec-fetch-dest': 'empty',
                'sec-fetch-mode': 'cors',
                'sec-fetch-site': 'same-origin',
                'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 YaBrowser/25.10.0.0 Safari/537.36',
                'x-queryid': 'qid826484063176061750920251125164101',
                'x-requested-with': 'XMLHttpRequest',
                'x-spa-version': '13.14.1',
                'x-userid': '0',
                # 'cookie': 'external-locale=ru; _wbauid=8264840631760617509; wbx-validation-key=0b839a54-4dfe-4b2a-a0a6-edf2ba0b6db4; x-supplier-id-external=0cd68178-a8ec-409f-8817-787cd8da5b16; _cp=1; x_wbaas_token=1.1000.8fe21158a99145ff97177af0a0a36575.MHw3Ny4yNDMuMjkuMTI2fE1vemlsbGEvNS4wIChXaW5kb3dzIE5UIDEwLjA7IFdpbjY0OyB4NjQpIEFwcGxlV2ViS2l0LzUzNy4zNiAoS0hUTUwsIGxpa2UgR2Vja28pIENocm9tZS8xNDAuMC4wLjAgWWFCcm93c2VyLzI1LjEwLjAuMCBTYWZhcmkvNTM3LjM2fDE3NjUyNjQ4MzN8cmV1c2FibGV8MnxleUpvWVhOb0lqb2lJbjA9fDB8M3wxNzY0NjYwMDMzfDE=.MEUCIQDdvJ69jwv5y4qD8HkfWPfE7NttWzAApDZDDXLrC9RQ3AIgWQJ8j9JsEFOgO9iAV1IKCisj99mKd+a6TQweuIPwpnI=; __zzatw-wb=MDA0dC0cTHtmcDhhDHEWTT17CT4VHThHKHIzd2UqP2YlaE5eJjVRP0FaW1Q4NmdBEXUmCQg3LGBwVxlRExpceEdXeiwfFXdwKFIKEF5CSmllbQwtUlFRS19/Dg4/aU5ZQ11wS3E6EmBWGB5CWgtMeFtLKRZHGzJhXkZpdRVSCxFeQnN0dCw9bFBjShVQdBAICFpNRXkmKk8NFF1BR19vG3siXyoIJGM1Xz9EaVhTMCpYQXt1J3Z+KmUzPGwjY0deJEZYUnwrGw1pN2wXPHVlLwkxLGJ5MVIvE0tsP0caRFpbQDsyVghDQE1HFF9BWncyUlFRS2EQR0lrZU5TQixmG3EVTQgNND1aciIPWzklWAgSPwsmIBd7ayhUCg5gQEdxbxt/Nl0cOWMRCxl+OmNdRkc3FSR7dSYKCTU3YnAvTCB7SykWRxsyYV5GaXUVTzo/YUVCdHsmbG1SGkRdH0wTSgktGhh0citWOj9jcXJyLSpBVxlRDxZhDhYYRRcje0I3Yhk4QhgvPV8/YngiD2lIYCVHVVJ+KBwWe3EmS3FPLH12X30beylOIA0lVBMhP05yIeFnyg==; cfidsw-wb=XsAr4MnOWjEwlWh4U96Tq+AFP3MrXypo1QoN28dMc9pZnWZAHDjH0LSKjAyYELUerfy2t40DHt6gh+ARWmZoUTCf+tOrRHCIzokKcqffeGBN7zBQ+wNC6vG1cRQY/0QpTeyzO2X+yBSLAvZbHoaxbQlUi1ODq9ReFSe7ewQ=',
            }

            params = {
                'ab_testing': [
                    'false',
                    'false',
                ],
                'appType': '1',
                'curr': 'rub',
                'dest': '-1257786',
                'hide_dtype': '11',
                'inheritFilters': 'false',
                'page': str(page),
                'query': keyword,
                'resultset': 'catalog',
                'sort': 'popular',
                'spp': '30',
                'suppressSpellcheck': 'false',
            }

            response = requests.get(
                'https://www.wildberries.ru/__internal/u-search/exactmatch/ru/common/v18/search',
                params=params,
                cookies=cookies,
                headers=headers,
            )
            data = response.json()
            if not data.get('products'):
                break  # если товаров больше нет — выходим из while


            #if len(response.json()) != 0:
            result = find_item(response, id_articul=259711529) #259711529 78035685
            if result:
                arr.append(str((page - 1) * 100 + result))
                break
            page += 1
    #запись в exel
    write_to_sheet_1(arr)

def main():
    get_resp()


if __name__ == '__main__':
    main()

